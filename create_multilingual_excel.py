#!/usr/bin/env python3
"""
Create Excel file with multilingual store data for testing
"""
import openpyxl
from openpyxl import Workbook

def create_multilingual_excel():
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Stores Multilingual"
    
    # Headers (standard fields + multilingual fields)
    headers = [
        'ownerFirstName', 'ownerLastName', 'storeName', 'storeName_ku', 'storeName_ar',
        'phone', 'email', 'logo', 'CoverPhoto', 'latitude', 'longitude', 
        'Address', 'Address_ku', 'Address_ar', 'zone_id', 'module_id',
        'MinimumOrderAmount', 'Comission', 'Tax', 'DeliveryTime',
        'MinimumDeliveryFee', 'PerKmDeliveryFee', 'MaximumDeliveryFee',
        'ScheduleOrder', 'Status', 'SelfDeliverySystem', 'Veg', 'NonVeg',
        'FreeDelivery', 'TakeAway', 'Delivery', 'ReviewsSection', 'PosSystem',
        'storeOpen', 'FeaturedStore'
    ]
    
    # Add headers to first row
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Test data with Kurdish and Arabic translations
    test_data = [
        [
            'Ahmad', 'Hassan', 'Kurdistan Restaurant', 'چێشتخانەی کوردستان', 'مطعم كردستان',
            '+9647501234567', 'ahmad.kurdistan@test.com', 'def.png', 'def.png', 36.1916, 44.0092,
            'Downtown Erbil', 'ناوەندی هەولێر', 'وسط أربيل', 1, 2,
            5000, 10, 5, '30-45 min', 2000, 500, 15000,
            'yes', 'active', 'active', 'yes', 'yes', 'no', 'yes', 'yes',
            'active', 'active', 'yes', 'no'
        ],
        [
            'Fatima', 'Ali', 'Hawler Food Corner', 'گۆشەی خواردنی هەولێر', 'ركن طعام أربيل',
            '+9647501234568', 'fatima.hawler@test.com', 'def.png', 'def.png', 36.2000, 44.0100,
            'Sami Abdul Rahman Park', 'پارکی سامی عەبدولڕەحمان', 'حديقة سامي عبد الرحمن', 1, 2,
            3000, 8, 3, '20-35 min', 1500, 400, 12000,
            'yes', 'active', 'inactive', 'yes', 'no', 'yes', 'yes', 'yes',
            'active', 'active', 'yes', 'yes'
        ],
        [
            'Omar', 'Rashid', 'Slemani Delights', 'تامەکانی سلێمانی', 'مأكولات السليمانية',
            '+9647501234569', 'omar.slemani@test.com', 'def.png', 'def.png', 35.5617, 45.4329,
            'Sulaymaniyah Center', 'ناوەندی سلێمانی', 'مركز السليمانية', 1, 2,
            4000, 12, 6, '25-40 min', 1800, 600, 14000,
            'no', 'active', 'active', 'no', 'yes', 'no', 'yes', 'yes',
            'active', 'inactive', 'yes', 'no'
        ]
    ]
    
    # Add data rows
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save the file
    filename = 'test_stores_multilingual.xlsx'
    wb.save(filename)
    print(f"✅ Created Excel file: {filename}")
    print(f"📊 Data includes {len(test_data)} stores with Kurdish and Arabic translations")
    print(f"🌍 Languages: English (default), Kurdish Sorani (ku), Arabic (ar)")
    
if __name__ == "__main__":
    create_multilingual_excel()